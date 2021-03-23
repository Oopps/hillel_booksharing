from django.contrib.auth.mixins import LoginRequiredMixin
from django.http import HttpResponse
from django.shortcuts import redirect, get_object_or_404
from django.urls import reverse_lazy
from django.views import View
from django.views.generic import CreateView, UpdateView, DeleteView, ListView, TemplateView

from books.forms import BookForm, AuthorForm
from books.models import Book, RequestBook
from books.models import Author

from openpyxl import Workbook
from datetime import datetime
from books import model_choises as mch
from django.contrib import messages


class Index(TemplateView):
    template_name = 'index.html'


class BookList(ListView):
    model = Book

    def get_queryset(self):
        return self.model.objects.exclude(user=self.request.user).select_related('author', 'category')


class MyBookList(ListView):
    model = Book

    def get_queryset(self):
        return self.model.objects.filter(user=self.request.user).select_related('author', 'category')


class MyRequestedBooks(LoginRequiredMixin, ListView):
    queryset = RequestBook.objects.all()

    def get_queryset(self):
        queryset = super().get_queryset()
        return queryset.filter(recipient=self.request.user)


class RequestedBooks(LoginRequiredMixin, ListView):
    queryset = RequestBook.objects.all()
    template_name = 'books/requestedbook_list.html'

    def get_queryset(self):
        queryset = super().get_queryset()
        return queryset.filter(book__user=self.request.user)


class BookCreate(LoginRequiredMixin, CreateView):
    model = Book
    success_url = reverse_lazy('my-book-list')
    form_class = BookForm

    def get_success_url(self):
        messages.add_message(self.request, messages.INFO, 'Book was created!')
        return super().get_success_url()


class BookUpdate(LoginRequiredMixin, UpdateView):
    model = Book
    success_url = reverse_lazy('my-book-list')
    form_class = BookForm

    def get_success_url(self):
        messages.add_message(self.request, messages.INFO, 'Book was updated!')
        return super().get_success_url()


class BookDelete(LoginRequiredMixin, DeleteView):
    model = Book
    success_url = reverse_lazy('my-book-list')

    def get_success_url(self):
        messages.add_message(self.request, messages.INFO, 'Book was deleted!')
        return super().get_success_url()


class AuthorList(ListView):
    queryset = Author.objects.all()


class AuthorCreate(LoginRequiredMixin, CreateView):
    model = Author
    success_url = reverse_lazy('author-list')
    form_class = AuthorForm

    def get_success_url(self):
        messages.add_message(self.request, messages.INFO, 'Author was created!')
        return super().get_success_url()


class AuthorUpdate(LoginRequiredMixin, UpdateView):
    model = Author
    success_url = reverse_lazy('author-list')
    form_class = AuthorForm

    def get_success_url(self):
        messages.add_message(self.request, messages.INFO, 'Author was updated!')
        return super().get_success_url()


class AuthorDelete(LoginRequiredMixin, DeleteView):
    model = Author
    success_url = reverse_lazy('author-list')

    def get_success_url(self):
        messages.add_message(self.request, messages.INFO, 'Author was deleted!')
        return super().get_success_url()


class RequestBookCreate(LoginRequiredMixin, View):

    def get(self, request, book_id):
        book = get_object_or_404(Book, pk=book_id)
        if not RequestBook.objects.filter(book=book, recipient=request.user, status=mch.STATUS_IN_PROGRESS).exists():
            RequestBook.objects.create(book=book, recipient=request.user, status=mch.STATUS_IN_PROGRESS)
        return redirect('book-list')


class _ChangeRequestBaseView(LoginRequiredMixin, View):
    CURRENT_STATUS = None
    NEW_STATUS = None
    REDIRECT_NAME = None
    MESSAGE = None

    def get(self, request, request_id):
        request_obj = get_object_or_404(RequestBook, pk=request_id, status=self.CURRENT_STATUS)
        request_obj.status = self.NEW_STATUS
        request_obj.save(update_fields=('status',))

        if self.MESSAGE:
            messages.add_message(request, messages.INFO, self.MESSAGE)

        return redirect(self.REDIRECT_NAME)


class RequestBookConfirm(_ChangeRequestBaseView):
    CURRENT_STATUS = mch.STATUS_IN_PROGRESS
    NEW_STATUS = mch.STATUS_CONFIRMED
    REDIRECT_NAME = 'requested-books'
    MESSAGE = 'Book request confirmed!'


class RequestBookSentViaEmail(_ChangeRequestBaseView):
    CURRENT_STATUS = mch.STATUS_CONFIRMED
    NEW_STATUS = mch.STATUS_SENT_TO_RECIPIENT
    REDIRECT_NAME = 'requested-books'
    MESSAGE = 'Book has been sent by mail!'


class RequestBookReceivedBook(_ChangeRequestBaseView):
    CURRENT_STATUS = mch.STATUS_SENT_TO_RECIPIENT
    NEW_STATUS = mch.STATUS_RECIPIENT_RECEIVED_BOOK
    REDIRECT_NAME = 'my-requested-books'


class RequestBookReturn(_ChangeRequestBaseView):
    CURRENT_STATUS = mch.STATUS_RECIPIENT_RECEIVED_BOOK
    NEW_STATUS = mch.STATUS_SENT_BACK_TO_OWNER
    REDIRECT_NAME = 'my-requested-books'


class RequestBookOwnerReceived(_ChangeRequestBaseView):
    CURRENT_STATUS = mch.STATUS_SENT_BACK_TO_OWNER
    NEW_STATUS = mch.STATUS_OWNER_RECEIVED_BACK
    REDIRECT_NAME = 'requested-books'


class RequestBookReject(_ChangeRequestBaseView):

    def get(self, request, request_id):
        request_obj = get_object_or_404(RequestBook, pk=request_id, status=mch.STATUS_IN_PROGRESS)
        request_obj.status = mch.STATUS_REJECTED
        request_obj.save(update_fields=('status', ))
        return redirect('requested-books')


def export_books_to_xlsx(request):
    """
    Downloads all books as Excel file with a single worksheet
    """
    book_queryset = Book.objects.all().select_related('author', 'category')

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = 'attachment; filename={date}-books.xlsx'.format(
        date=datetime.now().strftime('%Y-%m-%d'),
    )
    workbook = Workbook()

    # Get active worksheet/tab
    worksheet = workbook.active
    worksheet.title = 'Books'

    # Define the titles for columns
    columns = [
        'Title',
        'Publish_year',
        'Review',
        'Condition',
        'Author',
        'Category',
    ]
    row_num = 1

    # Assign the titles for each cell of the header
    for col_num, column_title in enumerate(columns, 1):
        cell = worksheet.cell(row=row_num, column=col_num)
        cell.value = column_title

    # Define the data for each cell in the row
    for book in book_queryset:
        row_num += 1
        row = [
            book.title,
            book.publish_year,
            book.review,
            book.condition,
            book.author.second_name,
            book.category,
        ]

        # Assign the data for each cell of the row
        for col_num, cell_value in enumerate(row, 1):
            cell = worksheet.cell(row=row_num, column=col_num)
            cell.value = cell_value

    workbook.save(response)
    return response
